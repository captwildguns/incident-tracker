"""
Generates Incident-Tracker-Seed-Catalog.xlsx — a hand-off catalog for the dev
to seed workflows, workflow steps, incident types, step templates, and email
templates into the database. Includes a SQL Guide sheet (DDL + mapping notes).

Source of truth (kept in sync manually with these files):
  - Figma files/src/components/incidents/IncidentTypes.ts
  - Figma files/src/data/workflows.ts
  - Figma files/src/components/workflows/WorkflowStepLibrary.tsx  (workflowStepTemplates)
  - Figma files/src/data/email-templates.ts
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────
# DATA
# ──────────────────────────────────────────────────────────────────────────

INCIDENT_TYPES = [
    # id, label, category, default_severity, applicable_to, description
    ("disruptive-behavior", "Disruptive Behavior", "Behavioral", "Low", "student",
     "Offensive language, excessive noise, harassment, bullying, refusal of driver directives, or any disruptive conduct on the bus"),
    ("safety-violation", "Safety Violation", "Safety", "Medium", "student",
     "Seat or seatbelt refusal, unsafe movement, window misuse, emergency exit misuse, wrong stop exit, or eating/drinking on the bus"),
    ("physical-altercation", "Physical Altercation", "Aggression / Violence", "High", "student",
     "Fighting, physical assault, throwing objects, or verbal/physical threats directed toward another student or any person on the bus"),
    ("property-damage", "Property Damage", "Property", "Medium", "student",
     "Vandalism or damage to the bus, equipment, or personal belongings requiring restitution"),
    ("weapon-prohibited-items", "Weapon / Prohibited Items", "Prohibited", "Critical", "student",
     "Possession of a weapon, weapon-like object, tobacco, vaping devices, illegal substances, or any other prohibited materials on the bus"),
    ("witness-bystander", "Witness / Bystander Statement", "Informational", "Low", "student",
     "Non-disciplinary record capturing the account of a student who witnessed or tried to help during another incident. Use this instead of adding a bystander to a disciplinary incident (e.g. a fight) so their record is not flagged for behavior they were not part of."),
]

# Workflows: id, name, description, incident_types[], severity_levels[]
# IDs are sequential, assigned alphabetically by workflow name.
WORKFLOWS = [
    ("WF-001", "Disruptive Behavior Response",
     "Workflow for all disruptive behavior incidents including offensive language, excessive noise, bullying, harassment, defiance toward the driver, and unauthorized device usage",
     ["Disruptive Behavior"], ["Low"]),
    ("WF-002", "Physical Altercation Response",
     "Workflow for handling physical altercations and threatening behavior between students requiring immediate intervention and parent notification",
     ["Physical Altercation"], ["High"]),
    ("WF-003", "Prohibited Items Response",
     "Workflow for handling possession of prohibited items including tobacco, harmful items, illegal substances, and inappropriate materials",
     ["Weapon / Prohibited Items"], ["Critical"]),
    ("WF-004", "Property Damage Investigation",
     "Workflow for investigating and resolving vandalism and property damage incidents",
     ["Property Damage"], ["Medium"]),
    ("WF-005", "Safety Violation Response",
     "Workflow for all student safety violations on the bus including seat/seatbelt refusal, unsafe movement, window misuse, emergency exit misuse, wrong stop exit, and eating/drinking",
     ["Safety Violation"], ["Medium"]),
    ("WF-006", "Witness / Bystander Statement",
     "Non-disciplinary workflow for capturing the account of a student who witnessed or tried to help during another incident. No disciplinary action is taken; the statement is recorded and linked to the related incident.",
     ["Witness / Bystander Statement"], ["Low"]),
]

# Workflow steps. Each: workflow_id, step_key, order, name, description, assigned_role,
# estimated_duration, required, requires_approval, approvers[], trigger_type,
# trigger_delay_amount, trigger_delay_unit, condition (field/op/value or ""),
# email_template, notify_on_start, notify_on_complete, notify_assignee, notify_approvers, additional_recipients[]
# Use None for "no email notifications configured".
# NOTE: every step is manually triggered and no longer references an email template.
WORKFLOW_STEPS = [
    # WF-001 Disruptive Behavior Response
    ("WF-001", "step-1", 1, "Driver Warning & Documentation",
     "Driver issues warning, logs incident details, and contacts dispatch if situation cannot be de-escalated",
     "Driver", "15 minutes", True, False, [], "manual", None, None, "", None, None, None, None, None, []),
    ("WF-001", "step-2", 2, "Parent Notification",
     "Notify parent/guardian of the incident and expected behavior standards",
     "Safety Coordinator", "10 minutes", True, False, [], "manual", None, None, "",
     None, False, True, True, False, []),
    ("WF-001", "step-3", 3, "Close Incident",
     "Review and close incident if no further action needed",
     "Safety Coordinator", "5 minutes", True, False, [], "manual", None, None, "", None, None, None, None, None, []),

    # WF-002 Physical Altercation Response
    ("WF-002", "step-1", 1, "Immediate Driver Response",
     "Safely stop bus, assess threat level, separate involved students, and contact dispatch/911 if there is imminent danger or a weapon",
     "Driver", "15 minutes", True, False, [], "manual", None, None, "", None, None, None, None, None, []),
    ("WF-002", "step-2", 2, "Submit Incident Report",
     "Driver submits the incident details — what occurred, injuries, and any threatening behavior — through the tablet app. Submitting the report notifies the safety coordinator.",
     "Driver", "15 minutes", True, False, [], "manual", None, None, "",
     None, None, None, None, None, []),
    ("WF-002", "step-3", 3, "Parent Notification",
     "Contact parents of all students involved and explain incident",
     "Safety Coordinator", "30 minutes", True, False, [], "manual", None, None, "",
     None, False, True, True, False, []),
    ("WF-002", "step-4", 4, "Disciplinary Action Review",
     "Administrator reviews incident and determines appropriate disciplinary measures",
     "Administrator", "1 hour", True, True, ["Administrator"],
     "manual", None, None, "", "Action Required", False, True, True, True, []),
    ("WF-002", "step-5", 5, "Documentation & Close",
     "Complete incident documentation and close case",
     "Safety Coordinator", "20 minutes", True, False, [], "manual", None, None, "", None, None, None, None, None, []),

    # WF-003 Prohibited Items Response
    ("WF-003", "step-1", 1, "Confiscation & Secure",
     "Driver safely confiscates item (if possible) and secures it; do not handle weapons or suspected drugs directly—contact dispatch",
     "Driver", "15 minutes", True, False, [], "manual", None, None, "", None, None, None, None, None, []),
    ("WF-003", "step-2", 2, "Submit Incident Report",
     "Driver submits the incident details — what occurred, the prohibited item involved, and how it was confiscated or secured — through the tablet app. Submitting the report notifies the safety coordinator.",
     "Driver", "15 minutes", True, False, [], "manual", None, None, "",
     None, None, None, None, None, []),
    ("WF-003", "step-3", 3, "Parent Notification",
     "Contact parent/guardian to inform them of the prohibited item and policy violation",
     "Safety Coordinator", "20 minutes", True, False, [], "manual", None, None, "",
     None, False, True, True, False, []),
    ("WF-003", "step-4", 4, "Disciplinary & Legal Action Review",
     "Determine appropriate disciplinary action per district policy; coordinate with law enforcement if applicable",
     "Administrator", "45 minutes", True, True, ["Administrator"],
     "manual", None, None, "", "Action Required", False, True, True, True, []),
    ("WF-003", "step-5", 5, "Documentation & Close",
     "Complete all documentation including evidence chain of custody if applicable, and close case",
     "Safety Coordinator", "20 minutes", True, False, [], "manual", None, None, "", None, None, None, None, None, []),

    # WF-004 Property Damage Investigation
    ("WF-004", "step-1", 1, "Damage Assessment & Photo Documentation",
     "Driver photographs damage",
     "Driver", "15 minutes", True, False, [], "manual", None, None, "", None, None, None, None, None, []),
    ("WF-004", "step-2", 2, "Fleet Manager Review",
     "Fleet manager assesses damage and provides repair estimate",
     "Fleet Manager", "30 minutes", True, False, [], "manual", None, None, "",
     "Action Required", False, True, True, False, []),
    ("WF-004", "step-3", 3, "Parent Notification & Restitution",
     "Contact parents and discuss restitution for damages",
     "Safety Coordinator", "45 minutes", True, False, [], "manual", None, None, "",
     None, False, True, True, False, []),
    ("WF-004", "step-4", 4, "Disciplinary Action",
     "Implement disciplinary measures per district policy",
     "School Principal", "20 minutes", True, True, ["Administrator"], "manual", None, None, "",
     "Action Required", False, True, True, True, []),
    ("WF-004", "step-5", 5, "Repair Scheduling",
     "Schedule and complete vehicle repairs",
     "Fleet Manager", "2 hours", True, False, [], "manual", None, None, "",
     "Action Required", False, True, True, False, []),
    ("WF-004", "step-6", 6, "Documentation & Close",
     "Complete incident documentation and close case",
     "Safety Coordinator", "20 minutes", True, False, [], "manual", None, None, "", None, None, None, None, None, []),

    # WF-005 Safety Violation Response
    ("WF-005", "step-1", 1, "Immediate Safety Response",
     "Driver addresses the safety issue, secures the situation, and documents the incident",
     "Driver", "15 minutes", True, False, [], "manual", None, None, "", None, None, None, None, None, []),
    ("WF-005", "step-2", 2, "Submit Incident Report",
     "Driver submits the incident details — what occurred, the safety violation involved, and any action taken to secure the situation — through the tablet app. Submitting the report notifies the safety coordinator.",
     "Driver", "15 minutes", True, False, [], "manual", None, None, "",
     None, None, None, None, None, []),
    ("WF-005", "step-3", 3, "Parent Notification",
     "Contact parent/guardian to inform them of the safety violation and reinforce bus safety expectations",
     "Safety Coordinator", "15 minutes", True, False, [], "manual", None, None, "",
     None, False, True, True, False, []),
    ("WF-005", "step-4", 4, "Disciplinary Action Review",
     "Administrator reviews incident and determines appropriate disciplinary measures",
     "Administrator", "1 hour", True, True, ["Administrator"], "manual", None, None,
     "", "Action Required", False, True, True, True, []),
    ("WF-005", "step-5", 5, "Documentation & Close",
     "Complete all documentation and close the incident",
     "Safety Coordinator", "10 minutes", True, False, [], "manual", None, None, "", None, None, None, None, None, []),

    # WF-006 Witness / Bystander Statement
    ("WF-006", "step-1", 1, "Record Witness Statement",
     "Driver or staff records the witness/bystander account of what they saw or how they helped. No fault is assigned to this student.",
     "Driver", "10 minutes", False, False, [], "manual", None, None, "", None, None, None, None, None, []),
]

# Roles to notify by email for a step, keyed by (workflow_id, step_key).
# Most steps notify no extra role; only the entries below set notify_groups.
NOTIFY_GROUPS = {
    ("WF-002", "step-4"): ["Safety Coordinator"],
    ("WF-003", "step-4"): ["Safety Coordinator"],
    ("WF-004", "step-2"): ["Safety Coordinator"],
    ("WF-004", "step-4"): ["Safety Coordinator"],
    ("WF-004", "step-5"): ["Safety Coordinator"],
    ("WF-005", "step-4"): ["Safety Coordinator"],
}

# Step templates: id, name, category, description, default_group, default_duration,
# requires_approval, notify_on_start, notify_on_complete, notify_assignee, tags[]
STEP_TEMPLATES = [
    ("comm-parent-notify", "Parent/Guardian Notification", "Notification",
     "Contact parent or guardian to inform them of the incident and expected next steps",
     "Safety Coordinator", "20 minutes", False, False, True, True,
     ["parent", "guardian", "notification", "contact"]),
    ("doc-photo-evidence", "Photo & Evidence Documentation", "Review & Action",
     "Photograph damage or relevant scene, document physical evidence",
     "Driver", "20 minutes", False, None, None, None,
     ["photo", "evidence", "documentation", "damage"]),
    ("admin-disciplinary-review", "Disciplinary Review", "Review & Action",
     "Administrator reviews incident and determines appropriate disciplinary action; requires approval before proceeding",
     "School Principal", "1 hour", True, False, True, True,
     ["disciplinary", "review", "administrator", "approval"]),
    ("admin-police-report", "Law Enforcement Contact", "Review & Action",
     "File police report or coordinate with law enforcement for criminal incidents",
     "Administrator", "1 hour", True, None, None, None,
     ["police", "law enforcement", "criminal", "report"]),
    ("follow-close-incident", "Documentation & Close", "Close Out",
     "Complete all incident documentation, finalize the record, and close the case",
     "Safety Coordinator", "15 minutes", False, None, None, None,
     ["close", "documentation", "finalize", "complete"]),
]

# Extra email config for step templates that mirror configured workflow steps,
# keyed by template id: notify_approvers, notify_groups (roles), email_template.
TEMPLATE_EMAIL_EXTRA = {
    "admin-disciplinary-review": {"notify_approvers": True, "notify_groups": ["Safety Coordinator"], "email_template": "Action Required"},
}

# Email templates: id, name, category, is_default, last_modified, description, subject, body, variables[]
# IDs are assigned sequentially in alphabetical order by template name.
EMAIL_TEMPLATES = [
    ("ET-001", "Action Required", "Notification", True, "2026-06-19",
     "General notification that a workflow step requires action by the assigned role. Suitable for routine and time-sensitive steps alike.",
     "Action Required: {{step_name}} - {{incident_id}}",
     "Hello,\n\nA workflow step requires your attention:\n\nStep: {{step_name}}\nIncident: {{incident_id}}\nSeverity: {{severity}}\nAssigned To: {{assigned_role}}\n\nPlease log in to the Incident Tracker to review and complete this step.\n\nThank you,\nIncident Tracker System",
     ["step_name", "incident_id", "severity", "assigned_role"]),
    ("ET-002", "Approval Request", "Approval", True, "2026-03-10",
     "Sent to designated approvers when a step requires approval before proceeding.",
     "Approval needed for workflow step: {{step_name}} ({{incident_id}})",
     "Hello,\n\nYour approval is needed for the following workflow step:\n\nStep: {{step_name}}\nIncident: {{incident_id}}\nRequested By: {{requested_by}}\nDate: {{request_date}}\n\nDetails:\n{{step_description}}\n\nPlease log in to the Incident Tracker to review and approve or reject this step.\n\nThank you,\nIncident Tracker System",
     ["step_name", "incident_id", "requested_by", "request_date", "step_description"]),
    ("ET-003", "Parent/Guardian Notification", "Notification", True, "2026-03-17",
     "Sent to parents or guardians when their child is involved in a bus incident. Used across most student-related workflows.",
     "Incident Involving Your Child - {{incident_type}} - {{incident_id}}",
     "Hello,\n\nWe are writing to inform you of an incident that occurred on {{incident_date}} involving your child, {{student_name}}, while riding {{bus_route}}.\n\nIncident Type: {{incident_type}}\nSeverity: {{severity}}\nLocation: {{incident_location}}\n\nSummary:\n{{incident_summary}}\n\nImmediate actions taken:\n{{actions_taken}}\n\nIf you have any questions or concerns, please contact INSERT NAME at PHONE# or EMAIL.\n\nSincerely,\n{{sender_name}}\n{{sender_title}}\nStudent Transportation Department",
     ["student_name", "bus_route", "incident_date", "incident_type", "severity",
      "incident_location", "incident_summary", "actions_taken", "sender_name", "sender_title"]),
]

# ──────────────────────────────────────────────────────────────────────────
# STYLING
# ──────────────────────────────────────────────────────────────────────────

BRAND_BLUE = "4A6FA5"
BRAND_OLIVE = "7B8458"
HEADER_FILL = PatternFill("solid", fgColor=BRAND_BLUE)
SECTION_FILL = PatternFill("solid", fgColor=BRAND_OLIVE)
ALT_FILL = PatternFill("solid", fgColor="EEF1F6")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
CELL_FONT = Font(name="Calibri", size=10)
MONO_FONT = Font(name="Consolas", size=10)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def b(val):
    """Render a boolean/None as DB-friendly text."""
    if val is None:
        return ""
    return "TRUE" if val else "FALSE"


def joinlist(val):
    return ", ".join(val) if val else ""


def write_table(ws, headers, rows, widths, wrap_cols=()):
    """Write a header row + data rows with styling, autofilter, frozen header."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = BORDER
    for ri, row in enumerate(rows, start=2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=c, value=val)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = WRAP_TOP if (c - 1) in wrap_cols else TOP
            if ri % 2 == 0:
                cell.fill = ALT_FILL
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    ws.row_dimensions[1].height = 28


# ──────────────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ──────────────────────────────────────────────────────────────────────────

wb = Workbook()

# ---- README ----
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 120

def line(text, *, title=False, head=False, mono=False):
    r = ws.max_row + 1 if ws.max_row > 1 or ws["B1"].value else 1
    cell = ws.cell(row=r, column=2, value=text)
    if title:
        cell.font = TITLE_FONT
        cell.fill = SECTION_FILL
        ws.row_dimensions[r].height = 24
    elif head:
        cell.font = Font(bold=True, size=11, color=BRAND_BLUE)
    elif mono:
        cell.font = MONO_FONT
    else:
        cell.font = CELL_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")

line("Incident Tracker — Database Seed Catalog", title=True)
line("")
line("Purpose", head=True)
line("This workbook catalogs the default configuration data for the Student Transportation Incident "
     "Tracker so it can be seeded into the database. It is the source-of-truth hand-off for the dev. "
     "All values come directly from the application's data files.")
line("")
line("Sheets", head=True)
line("• Incident Types     — the 6 incident categories an incident can be classified as.")
line("• Workflows          — the workflow templates (one per incident type). Header-level fields/defaults.")
line("• Workflow Steps     — the ordered steps for every workflow, with triggers, approval, and email config.")
line("• Step Templates     — the reusable step library used by the Workflow Builder to assemble new workflows.")
line("• Email Templates    — notification/approval email templates available in the system.")
line("")
line("Key relationships", head=True)
line("• Workflow  1───*  Workflow Steps           (workflow_id)")
line("• Workflow  *───*  Incident Type            (matched by the incident type LABEL, see 'incident_types' column)")
line("• Workflow  *───*  Severity Level           (a workflow applies to one or more severities)")
line("• Workflow Step  *───1  Email Template      (matched by the template NAME in 'email_template' column; blank = none)")
line("• Step Templates are standalone library entries; they are NOT instances and have no workflow_id.")
line("• List columns (incident_types, severity_levels, approvers, additional_recipients, tags, variables) hold "
     "comma-separated values; split them into one row each if normalizing into child tables.")
line("")
line("Notes for seeding", head=True)
line("• Booleans are written as TRUE / FALSE. Blank in a boolean column means 'not configured' (treat as NULL/false).")
line("• 'email_template' on a step references Email Templates.name. A blank means the step sends no templated email "
     "(though notify_* flags may still be set). Several steps reference the 'Action Required' template — the "
     "disciplinary-review step 4s (WF-002/003/004/005) and WF-004 steps 2 and 5; the other templates are "
     "seeded as a reusable library.")
line("• 'notify_groups' lists the roles that should receive the step's email (comma-separated). 'notify_on_start' / "
     "'notify_on_complete' indicate timing: notify before the step starts vs. after it completes.")
line("• Every workflow step is manually triggered (trigger_type = manual); there are no time-delay/conditional/"
     "auto-advance steps in the current configuration.")
line("• estimated_duration / default_duration are free-text labels ('15 minutes', '2 hours'); keep as text unless "
     "you parse them into minutes.")
line("• Email template variables are listed without the surrounding {{ }} braces; the body shows the {{placeholder}} form.")
line("• 'Witness / Bystander Statement' (incident type + WF-006) is intentionally non-disciplinary: no parent "
     "notification or approval steps, so a student who only witnessed/helped is not flagged on a disciplinary incident.")
line("• Severity values: Low, Medium, High, Critical.   applicable_to values: student, driver, both.")
line("")
line("Generated by gen_seed_catalog_xlsx.py — re-run after editing the app data files to refresh.")

# ---- Incident Types ----
ws = wb.create_sheet("Incident Types")
headers = ["id", "label", "category", "default_severity", "applicable_to", "description"]
rows = [list(t) for t in INCIDENT_TYPES]
write_table(ws, headers, rows, [24, 30, 22, 16, 14, 80], wrap_cols=(5,))

# ---- Workflows ----
ws = wb.create_sheet("Workflows")
headers = ["id", "name", "description", "incident_types", "severity_levels"]
rows = []
for wf in WORKFLOWS:
    wid, name, desc, itypes, sevs = wf
    rows.append([wid, name, desc, joinlist(itypes), joinlist(sevs)])
write_table(ws, headers, rows, [10, 32, 75, 28, 24], wrap_cols=(2,))

# ---- Workflow Steps ----
ws = wb.create_sheet("Workflow Steps")
headers = ["workflow_id", "step_key", "step_order", "name", "description", "assigned_role",
           "estimated_duration", "required", "requires_approval", "approvers",
           "trigger_type", "trigger_delay_amount", "trigger_delay_unit", "trigger_condition",
           "email_template", "notify_on_start", "notify_on_complete", "notify_assignee",
           "notify_approvers", "notify_groups", "additional_recipients"]
rows = []
for s in sorted(WORKFLOW_STEPS, key=lambda s: (s[0], s[2])):
    (wid, key, order, name, desc, role, dur, req, appr_req, approvers, ttype, damt, dunit,
     cond, etmpl, n_start, n_complete, n_assignee, n_appr, addl) = s
    rows.append([wid, key, order, name, desc, role, dur, b(req), b(appr_req), joinlist(approvers),
                 ttype, damt if damt is not None else "", dunit or "", cond,
                 etmpl or "", b(n_start), b(n_complete), b(n_assignee), b(n_appr),
                 joinlist(NOTIFY_GROUPS.get((wid, key), [])), joinlist(addl)])
write_table(ws, headers,
            rows,
            [12, 9, 10, 30, 60, 18, 15, 9, 14, 26, 14, 13, 13, 18, 26, 13, 15, 14, 14, 22, 26],
            wrap_cols=(4,))

# ---- Step Templates ----
ws = wb.create_sheet("Step Templates")
headers = ["id", "name", "category", "description", "default_group", "default_duration",
           "requires_approval", "notify_on_start", "notify_on_complete", "notify_assignee",
           "notify_approvers", "notify_groups", "email_template", "tags"]
rows = []
for t in STEP_TEMPLATES:
    (tid, name, cat, desc, grp, dur, appr, n_start, n_complete, n_assignee, tags) = t
    extra = TEMPLATE_EMAIL_EXTRA.get(tid, {})
    rows.append([tid, name, cat, desc, grp, dur, b(appr), b(n_start), b(n_complete), b(n_assignee),
                 b(extra.get("notify_approvers")), joinlist(extra.get("notify_groups", [])),
                 extra.get("email_template") or "", joinlist(tags)])
write_table(ws, headers, rows,
            [24, 32, 16, 60, 18, 16, 14, 14, 16, 14, 14, 22, 20, 36], wrap_cols=(3,))

# ---- Email Templates ----
ws = wb.create_sheet("Email Templates")
headers = ["id", "name", "category", "is_default", "description",
           "subject", "body", "variables"]
rows = []
for e in EMAIL_TEMPLATES:
    (eid, name, cat, default, mod, desc, subj, body, vars_) = e
    rows.append([eid, name, cat, b(default), desc, subj, body, joinlist(vars_)])
write_table(ws, headers, rows,
            [10, 28, 14, 11, 45, 50, 80, 45], wrap_cols=(4, 5, 6, 7))

# ──────────────────────────────────────────────────────────────────────────
OUT = "Incident-Tracker-Seed-Catalog.xlsx"
wb.save(OUT)
print("Wrote", OUT)
print("Sheets:", wb.sheetnames)
print("Workflows:", len(WORKFLOWS), "| Steps:", len(WORKFLOW_STEPS),
      "| Incident types:", len(INCIDENT_TYPES), "| Step templates:", len(STEP_TEMPLATES),
      "| Email templates:", len(EMAIL_TEMPLATES))
